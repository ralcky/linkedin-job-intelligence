"""
Generate professional portfolio screenshot
Shows Excel data in best light
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_showcase_excel():
    """Create visually appealing Excel for screenshot"""
    
    # Load data
    df = pd.read_json('output/SaaS_Jobs_CLEAN_FINAL.json')
    
    # Select best columns for display
    showcase_cols = [
        'job_title', 
        'company_name', 
        'location',
        'employee_count',
        'company_revenue_range',
        'tech_stack',
        'posted_date'
    ]
    
    df_display = df[showcase_cols].head(15)
    
    # Clean tech stack display
    df_display['tech_stack'] = df_display['tech_stack'].apply(
        lambda x: ', '.join(x[:5]) if isinstance(x, list) and x else 'N/A'
    )
    
    # Rename columns for clarity
    df_display.columns = [
        'Job Title',
        'Company',
        'Location',
        'Company Size',
        'Revenue Range',
        'Tech Stack',
        'Posted'
    ]
    
    # Save to Excel
    output_file = 'PORTFOLIO_SHOWCASE.xlsx'
    df_display.to_excel(output_file, index=False, sheet_name='VP+ SaaS Jobs')
    
    # Apply styling
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 12
    
    # Add title
    ws.insert_rows(1)
    ws['A1'] = 'LinkedIn VP+ Jobs Intelligence Dashboard'
    ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Freeze panes
    ws.freeze_panes = 'A3'
    
    # Add metadata footer
    last_row = ws.max_row + 2
    ws[f'A{last_row}'] = f'Dataset: {len(df)} total jobs | Generated: December 30, 2025 | Python + Playwright'
    ws[f'A{last_row}'].font = Font(italic=True, size=9, color="666666")
    
    wb.save(output_file)
    
    print(f"✅ Portfolio showcase created: {output_file}")
    print("\n📸 NEXT STEPS:")
    print("1. Open PORTFOLIO_SHOWCASE.xlsx in Excel")
    print("2. Zoom to 100%")
    print("3. Press Windows + Shift + S (screenshot tool)")
    print("4. Capture the Excel window")
    print("5. Save as: portfolio_screenshot.png")
    print("\n💡 TIP: Make sure first 10-12 rows are visible for best impact")

if __name__ == "__main__":
    create_showcase_excel()